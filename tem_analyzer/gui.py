import math
import os
import sys

from . import qt_bootstrap

qt_bootstrap.configure()  # must run before Qt is loaded

import cv2
import numpy as np
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QGroupBox, QFormLayout, QDoubleSpinBox, QSpinBox, QSplitter,
    QMessageBox, QHeaderView, QComboBox, QCheckBox, QProgressBar,
    QInputDialog, QScrollArea,
)
from PyQt5.QtCore import Qt, QPoint, QRectF
from PyQt5.QtGui import QImage, QPixmap, QPainter, QPen, QColor
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import openpyxl

from .analyzer import (
    ScaleBarDetector, ParticleAnalyzer, HAS_TESSERACT, load_image, save_image,
)
from . import config
from . import labels as labelstore
from datetime import datetime


class ImageLabel(QLabel):
    def __init__(self):
        super().__init__()
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumSize(400, 400)
        self.setStyleSheet("border: 1px solid #ccc; background: #222;")
        self._pixmap = None
        self.click_callback = None
        self.measure_callback = None
        self.measuring_callback = None
        self.measure_mode = False
        # Measurement endpoints live in original-image coordinates, not widget
        # ones, so the line stays on the same part of the image when the view
        # zooms or pans underneath it.
        self._drag_start = None
        self._drag_end = None
        self._anchor = None       # first end of a pending two-click measurement
        self.anchor_callback = None
        self.zoom_callback = None
        self._zoom = 1.0          # 1.0 = fit the whole image in the widget
        self._pan = QPoint(0, 0)  # offset from centred, in widget pixels
        self._pan_from = None
        self._panned = False
        # Pixmap pixels per original-image pixel, so callbacks can report
        # coordinates in the only unit the measurements are in.
        self.source_scale = 1.0

    MAX_ZOOM = 16.0

    def set_image(self, pixmap, reset_view=True, source_scale=1.0):
        self._pixmap = pixmap
        self.source_scale = source_scale or 1.0
        self._drag_start = self._drag_end = self._anchor = None
        if reset_view:
            self._zoom = 1.0
            self._pan = QPoint(0, 0)
        self._apply_view()

    def _fit_size(self):
        """Size the image would take scaled to fit the widget."""
        return self._pixmap.size().scaled(self.size(), Qt.KeepAspectRatio)

    def _scaled_size(self):
        """On-screen size of the whole image at the current zoom.

        Pure arithmetic: the geometry is needed on every mouse move, and
        actually resampling the pixmap to ask how big it is costs seconds and
        hundreds of megabytes once zoomed in on a 2048px image.
        """
        return self._fit_size() * self._zoom

    def zoom_percent(self):
        if not self._pixmap or self._pixmap.width() == 0:
            return 100.0
        return self._scaled_size().width() / self._pixmap.width() * 100.0

    def _apply_view(self):
        """Single epilogue for anything that changes what is on screen."""
        if self._pixmap:
            self._clamp_pan()
        self.update()
        if self.zoom_callback:
            self.zoom_callback(self.zoom_percent())

    def set_zoom(self, zoom, anchor=None):
        """Zoom about `anchor` (widget coords), keeping that point in place."""
        if not self._pixmap:
            return
        zoom = min(max(float(zoom), 1.0), self.MAX_ZOOM)
        if abs(zoom - self._zoom) < 1e-6:
            return
        if anchor is None:
            anchor = QPoint(self.width() // 2, self.height() // 2)

        before = self._offset()
        ratio = zoom / self._zoom
        self._zoom = zoom
        # The image point under the anchor is (anchor - offset) / scale; hold
        # it fixed by moving the pan so that quantity is unchanged.
        rel = anchor - before
        self._pan = anchor - self._centre_offset() - QPoint(int(rel.x() * ratio),
                                                            int(rel.y() * ratio))
        self._apply_view()

    def step_zoom(self, factor):
        self.set_zoom(self._zoom * factor)

    def reset_view(self):
        self._zoom = 1.0
        self._pan = QPoint(0, 0)
        self._apply_view()

    def _clamp_pan(self):
        """Keep the image covering the widget; centre it on the axes it can't."""
        scaled = self._scaled_size()
        slack_x = max(0, (scaled.width() - self.width()) // 2)
        slack_y = max(0, (scaled.height() - self.height()) // 2)
        self._pan = QPoint(min(max(self._pan.x(), -slack_x), slack_x),
                           min(max(self._pan.y(), -slack_y), slack_y))

    def paintEvent(self, event):
        super().paintEvent(event)  # stylesheet background and border
        if not self._pixmap:
            return
        painter = QPainter(self)
        painter.setClipRect(self.contentsRect())

        scaled = self._scaled_size()
        off = self._offset()
        target = QRectF(off.x(), off.y(), scaled.width(), scaled.height())
        visible = target.intersected(QRectF(self.rect()))
        if not visible.isEmpty():
            # Draw only the part of the image the widget can show. Scaling the
            # whole pixmap instead would cost time and memory proportional to
            # the zoom (655 MB at 16x on a 2048px image) to throw nearly all
            # of it away.
            fx = self._pixmap.width() / target.width()
            fy = self._pixmap.height() / target.height()
            source = QRectF((visible.x() - target.x()) * fx,
                            (visible.y() - target.y()) * fy,
                            visible.width() * fx, visible.height() * fy)
            # Smoothing averages neighbouring pixels away; past ~1.5x that
            # blurs exactly the grain and edges one zooms in to look at.
            painter.setRenderHint(QPainter.SmoothPixmapTransform,
                                  target.width() <= self._pixmap.width() * 1.5)
            painter.drawPixmap(visible, self._pixmap, source)

        painter.setPen(QPen(QColor(255, 60, 60), 2))
        if self._drag_start and self._drag_end:
            a, b = self._to_widget(self._drag_start), self._to_widget(self._drag_end)
            painter.drawLine(a, b)
            for end in (a, b):
                painter.drawLine(end.x(), end.y() - 6, end.x(), end.y() + 6)
        elif self._anchor:
            # A placed first end, waiting for the second one; the user may zoom
            # and pan in between, so mark where it is.
            a = self._to_widget(self._anchor)
            painter.drawLine(a.x(), a.y() - 8, a.x(), a.y() + 8)
            painter.drawLine(a.x() - 8, a.y(), a.x() + 8, a.y())

    def _centre_offset(self):
        scaled = self._scaled_size()
        return QPoint(int((self.width() - scaled.width()) / 2),
                      int((self.height() - scaled.height()) / 2))

    def _offset(self):
        return self._centre_offset() + self._pan

    def _to_image_coords(self, pos, clamp=False):
        """Widget point -> original-image pixels.

        Returns None for a point outside the image unless `clamp` is set, which
        pulls it to the nearest edge instead - a measurement drag that overshoots
        the image by a few pixels should still measure, not vanish.
        """
        scaled = self._scaled_size()
        off = self._offset()
        px, py = pos.x() - off.x(), pos.y() - off.y()
        if not (0 <= px < scaled.width() and 0 <= py < scaled.height()):
            if not clamp:
                return None
            px = min(max(px, 0), scaled.width() - 1)
            py = min(max(py, 0), scaled.height() - 1)
        return (px * self._pixmap.width() / scaled.width() / self.source_scale,
                py * self._pixmap.height() / scaled.height() / self.source_scale)

    def _to_widget(self, point):
        """Original-image pixels -> widget point (inverse of _to_image_coords)."""
        scaled = self._scaled_size()
        off = self._offset()
        return QPoint(
            int(off.x() + point[0] * self.source_scale * scaled.width() / self._pixmap.width()),
            int(off.y() + point[1] * self.source_scale * scaled.height() / self._pixmap.height()))

    def resizeEvent(self, event):
        self._apply_view()
        super().resizeEvent(event)

    def wheelEvent(self, event):
        if not self._pixmap:
            return super().wheelEvent(event)
        steps = event.angleDelta().y() / 120.0
        if steps:
            self.set_zoom(self._zoom * (1.25 ** steps), event.pos())
            event.accept()
            return
        super().wheelEvent(event)

    def _constrain(self, pos, modifiers):
        """Hold Shift to lock the drag to the horizontal or vertical axis.

        Scale bars are drawn straight, so a hand-drawn line across one is
        always a little tilted - which both looks wrong and shortens the
        reading by the cosine of the angle.
        """
        if not (modifiers & Qt.ShiftModifier) or self._drag_start is None:
            return pos
        origin = self._to_widget(self._drag_start)
        dx = pos.x() - origin.x()
        dy = pos.y() - origin.y()
        if abs(dx) >= abs(dy):
            return QPoint(pos.x(), origin.y())
        return QPoint(origin.x(), pos.y())

    def _pan_button(self, event):
        """Middle-drag always pans; left-drag pans when not measuring."""
        return (event.buttons() & Qt.MiddleButton
                or (event.buttons() & Qt.LeftButton and not self.measure_mode))

    def clear_measurement(self, line=True):
        """Drop a pending first end, and the finished line unless asked to keep it."""
        self._anchor = None
        if line:
            self._drag_start = self._drag_end = None
        if self.anchor_callback:
            self.anchor_callback(False)
        self.update()

    def mousePressEvent(self, event):
        if not self._pixmap:
            return super().mousePressEvent(event)
        self._panned = False
        if self.measure_mode and event.button() == Qt.LeftButton:
            point = self._to_image_coords(event.pos(), clamp=True)
            # A pending anchor means this press is the far end of a two-click
            # measurement, so the line runs from there rather than from here.
            self._drag_start = self._anchor or point
            self._drag_end = point
            self.update()
        else:
            self._pan_from = (event.pos(), QPoint(self._pan))
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self.measure_mode and self._drag_start:
            self._drag_end = self._to_image_coords(
                self._constrain(event.pos(), event.modifiers()), clamp=True)
            self.update()
            if self.measuring_callback:
                self.measuring_callback(self._drag_start, self._drag_end,
                                        bool(event.modifiers() & Qt.ShiftModifier))
        elif self._pan_from and self._pan_button(event):
            origin, base = self._pan_from
            delta = event.pos() - origin
            if not self._panned and delta.manhattanLength() > 3:
                self._panned = True
                self.setCursor(Qt.ClosedHandCursor)
            if self._panned:
                self._pan = base + delta
                self._clamp_pan()
                self.update()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self.measure_mode and self._drag_start:
            self._drag_end = self._to_image_coords(
                self._constrain(event.pos(), event.modifiers()), clamp=True)
            start, end = self._drag_start, self._drag_end
            if math.hypot(end[0] - start[0], end[1] - start[1]) < 2:
                # Too short to be a drag. Treat it as placing one end: the user
                # can now zoom and pan to the other end and click again, which
                # is the only way to measure a bar wider than the window.
                self._anchor = start
                self._drag_start = self._drag_end = None
                if self.anchor_callback:
                    self.anchor_callback(True)
            else:
                self._anchor = None
                if self.anchor_callback:
                    self.anchor_callback(False)
                if self.measure_callback:
                    self.measure_callback(start, end)
            self.update()
        elif self._pan_from:
            # A press that didn't move the view is a click, not a pan, so the
            # core toggle still works while the image is zoomed in.
            if not self._panned and event.button() == Qt.LeftButton and self.click_callback:
                coords = self._to_image_coords(event.pos())
                if coords:
                    self.click_callback(*coords)
            self.setCursor(Qt.CrossCursor if self.measure_mode else Qt.ArrowCursor)
        self._pan_from = None
        self._panned = False
        super().mouseReleaseEvent(event)


class HistogramCanvas(FigureCanvas):
    def __init__(self, parent=None):
        self.fig = Figure(figsize=(4, 3), dpi=100)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setMinimumHeight(250)

    def plot(self, diameters, unit="nm"):
        self.ax.clear()
        if not diameters:
            self.draw()
            return

        n_bins = max(5, min(30, len(diameters) // 3))
        self.ax.hist(diameters, bins=n_bins, color="#4A90D9", edgecolor="#2C5F8A", alpha=0.85)
        self.ax.set_xlabel(f"Diameter ({unit})", fontsize=10)
        self.ax.set_ylabel("Count", fontsize=10)
        self.ax.set_title("Particle Size Distribution", fontsize=11, fontweight="bold")

        mean_val = np.mean(diameters)
        self.ax.axvline(mean_val, color="#E74C3C", linestyle="--", linewidth=1.5,
                        label=f"Mean: {mean_val:.1f} {unit}")
        self.ax.legend(fontsize=9)
        self.fig.tight_layout()
        self.draw()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("TEM Particle Analyzer")
        self.setMinimumSize(1200, 800)
        self.image = None
        self.original_image = None
        self.particles = []
        self.result_image = None
        self.nm_per_px = None
        self.scale_text = None
        self.unit = "nm"
        self._measuring_shown = None
        #: Hand-measurement calibration. When calibrating, a drag across a
        #: particle records its diameter; applying them fixes where the edge is
        #: for the whole field, and the place is remembered so the next image
        #: of the same specimen reuses it.
        self._calib_mode = False
        self._calib_refs = []
        self.wall_place = None
        #: Ground-truth archive: hand-measured true diameters, per image, bound
        #: to positions rather than particle numbers. Purely data - it never
        #: changes what the analyzer measures; it records what is true so the
        #: measurement can be scored against it.
        self._label_mode = False
        self._labels = labelstore.load(labelstore.default_path())
        self._image_path = None
        self._image_key = None
        #: The blurred image + field wall fraction the wall reader needs, built
        #: once per analysis and reused for every label click; cleared when the
        #: field is re-analysed.
        self._wall_ctx = None
        self._wall_analyzer = None

        self._build_ui()
        # Carry the last session's settings - above all a hand calibration -
        # into this one. A missing or broken file is ignored (load returns {}),
        # so a first run just keeps the defaults.
        loaded = self._load_and_apply_settings(config.default_path(), quiet=True)
        if loaded and self.wall_place is not None:
            self.statusBar().showMessage(
                f"이전 보정을 불러왔습니다 (벽 위치 {self.wall_place:.2f}). "
                "이미지를 로드해주세요.")
        else:
            self.statusBar().showMessage("이미지를 로드해주세요.")

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)

        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)

        left = QWidget()
        left_layout = QVBoxLayout(left)

        btn_layout = QHBoxLayout()
        self.btn_load = QPushButton("이미지 열기")
        self.btn_load.setMinimumHeight(36)
        self.btn_load.clicked.connect(self._load_image)
        btn_layout.addWidget(self.btn_load)

        self.btn_analyze = QPushButton("분석 실행")
        self.btn_analyze.setMinimumHeight(36)
        self.btn_analyze.setEnabled(False)
        self.btn_analyze.clicked.connect(self._run_analysis)
        btn_layout.addWidget(self.btn_analyze)

        # A dense field takes tens of seconds; without this the window looks
        # frozen. It is hidden until an analysis starts and again when it ends.
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setTextVisible(True)
        self.progress.setVisible(False)

        self.btn_export = QPushButton("Excel 내보내기")
        self.btn_export.setMinimumHeight(36)
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self._export_excel)
        btn_layout.addWidget(self.btn_export)

        self.btn_save_image = QPushButton("결과 이미지 저장")
        self.btn_save_image.setMinimumHeight(36)
        self.btn_save_image.setEnabled(False)
        self.btn_save_image.clicked.connect(self._save_result_image)
        btn_layout.addWidget(self.btn_save_image)
        left_layout.addLayout(btn_layout)
        left_layout.addWidget(self.progress)

        self.image_label = ImageLabel()
        self.image_label.click_callback = self._on_image_click
        self.image_label.measure_callback = self._on_measure_drag
        self.image_label.measuring_callback = self._on_scalebar_measuring
        self.image_label.anchor_callback = self._on_measure_anchor
        self.image_label.zoom_callback = self._on_zoom_changed
        left_layout.addWidget(self.image_label, stretch=3)

        zoom_layout = QHBoxLayout()
        zoom_layout.addStretch()
        for text, tip, slot in (
            ("−", "축소", lambda: self.image_label.step_zoom(1 / 1.5)),
            ("+", "확대", lambda: self.image_label.step_zoom(1.5)),
            ("맞춤", "창 크기에 맞추기", self.image_label.reset_view),
        ):
            btn = QPushButton(text)
            btn.setToolTip(tip)
            btn.setFixedWidth(60 if len(text) > 1 else 36)
            btn.clicked.connect(slot)
            zoom_layout.addWidget(btn)
        self.lbl_zoom = QLabel("100%")
        self.lbl_zoom.setMinimumWidth(60)
        self.lbl_zoom.setAlignment(Qt.AlignCenter)
        self.lbl_zoom.setToolTip("마우스 휠로 확대/축소, 드래그로 이동")
        zoom_layout.addWidget(self.lbl_zoom)
        left_layout.addLayout(zoom_layout)

        self.histogram = HistogramCanvas()
        left_layout.addWidget(self.histogram, stretch=1)

        right = QWidget()
        right_layout = QVBoxLayout(right)

        scale_group = QGroupBox("스케일바 설정")
        scale_form = QFormLayout()

        if HAS_TESSERACT:
            self.scale_status = QLabel("자동 감지 대기 중")
        else:
            self.scale_status = QLabel("OCR 미설치 - 수동 입력 모드")
        scale_form.addRow("상태:", self.scale_status)

        self.chk_manual = QCheckBox("수동 입력 사용")
        scale_form.addRow(self.chk_manual)

        self.spin_bar_px = QDoubleSpinBox()
        self.spin_bar_px.setRange(1, 10000)
        self.spin_bar_px.setValue(100)
        self.spin_bar_px.setEnabled(False)
        scale_form.addRow("스케일바 길이 (px):", self.spin_bar_px)

        self.btn_measure = QPushButton("이미지에서 스케일바 재기")
        self.btn_measure.setCheckable(True)
        self.btn_measure.setEnabled(False)
        self.btn_measure.setToolTip(
            "누른 뒤 이미지 위의 스케일바 양 끝을 드래그하면 픽셀 길이가 자동 입력됩니다.\n"
            "Shift를 누른 채 끌면 선이 수평/수직으로 고정되어 기울지 않습니다.")
        self.btn_measure.toggled.connect(self._toggle_measure_mode)
        scale_form.addRow(self.btn_measure)

        self.spin_bar_real = QDoubleSpinBox()
        self.spin_bar_real.setRange(0.01, 100000)
        self.spin_bar_real.setValue(100)
        self.spin_bar_real.setEnabled(False)
        scale_form.addRow("실제 길이:", self.spin_bar_real)

        self.combo_unit = QComboBox()
        self.combo_unit.addItems(["nm", "μm", "mm"])
        self.combo_unit.setEnabled(False)
        scale_form.addRow("단위:", self.combo_unit)

        # Connect only once the widgets the handler touches exist: checking the
        # box emits toggled immediately, and without OCR that fired before the
        # spin boxes were built.
        self.chk_manual.toggled.connect(self._toggle_manual_scale)
        if not HAS_TESSERACT:
            self.chk_manual.setChecked(True)

        scale_group.setLayout(scale_form)
        right_layout.addWidget(scale_group)

        param_group = QGroupBox("분석 파라미터")
        param_form = QFormLayout()

        self.spin_min_area = QSpinBox()
        self.spin_min_area.setRange(10, 10000)
        self.spin_min_area.setValue(100)
        param_form.addRow("최소 면적 (px):", self.spin_min_area)

        self.spin_max_area = QSpinBox()
        self.spin_max_area.setRange(0, 1000000)
        self.spin_max_area.setValue(0)
        self.spin_max_area.setSpecialValueText("제한 없음")
        param_form.addRow("최대 면적 (px):", self.spin_max_area)

        self.spin_circularity = QDoubleSpinBox()
        self.spin_circularity.setRange(0.1, 1.0)
        self.spin_circularity.setValue(0.5)
        self.spin_circularity.setSingleStep(0.05)
        param_form.addRow("원형도 필터:", self.spin_circularity)

        self.chk_watershed = QCheckBox("겹침 분리 (Watershed)")
        self.chk_watershed.setChecked(True)
        param_form.addRow(self.chk_watershed)

        self.chk_hollow = QCheckBox("Hollow 입자 모드")
        self.chk_hollow.setChecked(True)
        self.chk_hollow.setToolTip(
            "속이 빈 입자(실리카 등)의 링 형태를 채워서 검출합니다.\n"
            "일반 입자에서도 손해가 없으므로 켜 두어도 됩니다.")
        param_form.addRow(self.chk_hollow)

        self.chk_shell = QCheckBox("쉘 두께 / 공극률 측정")
        self.chk_shell.setChecked(False)
        self.chk_shell.setToolTip(
            "중공 입자의 안쪽 공동 경계를 찾아 쉘 두께와 공극률을 계산합니다.\n"
            "공동이 뚜렷하지 않은 입자는 측정하지 않습니다.")
        param_form.addRow(self.chk_shell)

        self.chk_core = QCheckBox("코어 검출 (Yolk-shell)")
        self.chk_core.setChecked(False)
        self.chk_core.setToolTip("중공 입자 내부의 코어 입자를 감지하고 보유 비율을 계산")
        param_form.addRow(self.chk_core)

        self.chk_sphere_edge = QCheckBox("구(球) 모서리 보정")
        self.chk_sphere_edge.setChecked(False)
        self.chk_sphere_edge.setToolTip(
            "구는 가장자리로 갈수록 두께가 0으로 얇아지므로, 아무리 초점을 맞춰도\n"
            "경계가 '계단'이 아니라 서서히 사라지는 모양입니다. 밝기 기준으로는\n"
            "이 지점을 잡을 수 없어 실제보다 1픽셀쯤 작게 나옵니다.\n"
            "이 옵션은 대비의 제곱이 반지름에 대해 직선이 되는 성질을 이용해\n"
            "그 직선이 0을 지나는 곳을 외경으로 삼습니다.\n\n"
            "구·중공 입자에서는 오차가 -1.6~-3.1%에서 ±0.9% 이내로 줄어듭니다.\n"
            "다만 두께가 일정한 원반형(계단 경계) 입자는 1픽셀쯤 크게 나옵니다.\n"
            "실제 TEM 입자는 구이므로 대개 켜는 쪽이 맞습니다.")
        param_form.addRow(self.chk_sphere_edge)

        self.chk_edge_auto = QCheckBox("경계 자동 판단")
        self.chk_edge_auto.setChecked(True)
        self.chk_edge_auto.setToolTip(
            "쉘 벽마다 바깥 지름의 위치를 따로 판단합니다 (권장).\n"
            "이웃과 맞닿은 쪽은 두 벽이 겹쳐 보이므로 그 한가운데가 입자 표면이고,\n"
            "빈 공간을 향한 쪽은 자기 벽 하나뿐이므로 그 바깥 끝이 표면입니다.\n"
            "끄면 아래 값이 모든 방향에 똑같이 적용됩니다.")
        param_form.addRow(self.chk_edge_auto)

        self.spin_edge = QSpinBox()
        self.spin_edge.setRange(0, 100)
        self.spin_edge.setValue(int(ParticleAnalyzer.DEFAULT_EDGE_LEVEL * 100))
        self.spin_edge.setSuffix(" %")
        self.spin_edge.setEnabled(False)
        self.spin_edge.setToolTip(
            "'경계 자동 판단'을 끈 경우에만 쓰입니다.\n"
            "테두리(링) 위 어디를 경계로 볼지 정합니다.\n"
            "0% = 링의 안쪽 가장자리, 50% = 링 한가운데, 100% = 링의 바깥 가장자리.\n"
            "입자 크기는 외경이므로 기본값은 바깥 가장자리(95%)입니다.\n"
            "손으로 재던 값과 어긋나면 이 값으로 맞추세요.")
        self.chk_edge_auto.toggled.connect(
            lambda on: self.spin_edge.setEnabled(not on))
        param_form.addRow("경계 기준:", self.spin_edge)

        self.chk_mark_inferred = QCheckBox("미측정 구간 가늘게 표시")
        self.chk_mark_inferred.setChecked(False)
        self.chk_mark_inferred.setToolTip(
            "입자끼리 맞닿아 경계가 보이지 않는 구간을 가는 선으로 구분해 표시합니다.\n"
            "어디까지가 실측인지 확인할 때 켜세요. 끄면 경계선을 한 굵기로 그립니다.")
        self.chk_mark_inferred.toggled.connect(self._redraw_results)
        param_form.addRow(self.chk_mark_inferred)

        param_group.setLayout(param_form)
        right_layout.addWidget(param_group)

        calib_group = QGroupBox("경계 보정 (직접 재기)")
        calib_form = QFormLayout()
        self.btn_calib = QPushButton("지름 재기 시작")
        self.btn_calib.setCheckable(True)
        self.btn_calib.setEnabled(False)
        self.btn_calib.setToolTip(
            "흐릿한 쉘 벽 위 어디가 '입자 표면'인지는 보는 사람마다 다릅니다.\n"
            "이 버튼을 누른 뒤 입자 몇 개의 지름을 가로질러 드래그하면,\n"
            "그 판단을 기준으로 전체 입자의 지름을 다시 맞춥니다.\n"
            "한 번 보정하면 같은 시료의 다른 배율 이미지에도 그대로 적용됩니다.")
        self.btn_calib.toggled.connect(self._toggle_calib_mode)
        calib_form.addRow(self.btn_calib)

        self.lbl_calib = QLabel("측정 0개")
        calib_form.addRow("수집:", self.lbl_calib)

        calib_btns = QHBoxLayout()
        self.btn_calib_apply = QPushButton("보정 적용")
        self.btn_calib_apply.setEnabled(False)
        self.btn_calib_apply.clicked.connect(self._apply_calibration)
        calib_btns.addWidget(self.btn_calib_apply)
        self.btn_calib_reset = QPushButton("초기화")
        self.btn_calib_reset.setEnabled(False)
        self.btn_calib_reset.clicked.connect(self._reset_calibration)
        calib_btns.addWidget(self.btn_calib_reset)
        calib_form.addRow(calib_btns)

        # Persist a calibration between launches. It saves to a file on apply
        # and loads it back on startup, so the boundary the user fixed once
        # keeps applying; the explicit buttons export or import a named file so
        # a different specimen can have its own saved calibration.
        settings_btns = QHBoxLayout()
        self.btn_settings_save = QPushButton("설정 저장…")
        self.btn_settings_save.setToolTip(
            "현재 보정값과 분석 파라미터를 파일로 저장합니다 (배율은 제외).")
        self.btn_settings_save.clicked.connect(self._save_settings_as)
        settings_btns.addWidget(self.btn_settings_save)
        self.btn_settings_load = QPushButton("설정 불러오기…")
        self.btn_settings_load.setToolTip(
            "저장한 보정·파라미터 파일을 불러옵니다.")
        self.btn_settings_load.clicked.connect(self._load_settings_from)
        settings_btns.addWidget(self.btn_settings_load)
        calib_form.addRow(settings_btns)

        calib_group.setLayout(calib_form)
        right_layout.addWidget(calib_group)

        label_group = QGroupBox("참값 라벨링 (데이터 기록)")
        label_form = QFormLayout()
        self.btn_label = QPushButton("라벨 클릭 시작")
        self.btn_label.setCheckable(True)
        self.btn_label.setEnabled(False)
        self.btn_label.setToolTip(
            "입자를 클릭해 그 자리에 '참 지름'을 기록합니다. 측정값은 바뀌지 않고,\n"
            "참값만 데이터로 쌓입니다(이미지+위치에 묶여 재분석해도 유지).\n"
            "이미 라벨한 곳을 다시 클릭하면 수정, 0을 입력하면 삭제됩니다.")
        self.btn_label.toggled.connect(self._toggle_label_mode)
        label_form.addRow(self.btn_label)

        self.lbl_label_count = QLabel("이 이미지 라벨 0개")
        label_form.addRow("기록:", self.lbl_label_count)

        self.btn_label_calib = QPushButton("이 라벨로 보정")
        self.btn_label_calib.setEnabled(False)
        self.btn_label_calib.setToolTip(
            "이 이미지에 기록한 참값을 기준으로 전체 입자의 지름을 다시 잡습니다.\n"
            "라벨은 이미 저장돼 있으므로, 여기서 나온 보정값도 저장되어\n"
            "같은 시료의 다음 이미지에 자동 적용됩니다 — 쌓을수록 라벨이 덜 필요해집니다.")
        self.btn_label_calib.clicked.connect(self._calibrate_from_labels)
        label_form.addRow(self.btn_label_calib)

        label_btns = QHBoxLayout()
        self.btn_label_report = QPushButton("정확도 리포트")
        self.btn_label_report.setEnabled(False)
        self.btn_label_report.clicked.connect(self._accuracy_report)
        label_btns.addWidget(self.btn_label_report)
        self.btn_label_export = QPushButton("CSV 내보내기")
        self.btn_label_export.setEnabled(False)
        self.btn_label_export.clicked.connect(self._export_labels_csv)
        label_btns.addWidget(self.btn_label_export)
        label_form.addRow(label_btns)

        label_group.setLayout(label_form)
        right_layout.addWidget(label_group)

        stats_group = QGroupBox("통계 결과")
        stats_form = QFormLayout()
        self.lbl_count = QLabel("-")
        self.lbl_mean = QLabel("-")
        self.lbl_std = QLabel("-")
        self.lbl_min = QLabel("-")
        self.lbl_max = QLabel("-")
        self.lbl_d10 = QLabel("-")
        self.lbl_d50 = QLabel("-")
        self.lbl_d90 = QLabel("-")
        self.lbl_core = QLabel("-")
        self.lbl_shell = QLabel("-")
        self.lbl_defect = QLabel("-")
        self.lbl_irregular = QLabel("-")
        self.lbl_porosity = QLabel("-")
        stats_form.addRow("입자 수:", self.lbl_count)
        stats_form.addRow("평균 직경:", self.lbl_mean)
        stats_form.addRow("표준편차:", self.lbl_std)
        stats_form.addRow("최소:", self.lbl_min)
        stats_form.addRow("최대:", self.lbl_max)
        stats_form.addRow("D10:", self.lbl_d10)
        stats_form.addRow("D50:", self.lbl_d50)
        stats_form.addRow("D90:", self.lbl_d90)
        stats_form.addRow("코어 보유율:", self.lbl_core)
        stats_form.addRow("쉘 두께:", self.lbl_shell)
        stats_form.addRow("불량 (코어잔존/파손):", self.lbl_defect)
        stats_form.addRow("비원형:", self.lbl_irregular)
        stats_form.addRow("공극률:", self.lbl_porosity)
        stats_group.setLayout(stats_form)
        right_layout.addWidget(stats_group)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["#", "직경", "면적"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSortingEnabled(True)
        self.table.setMinimumHeight(180)
        right_layout.addWidget(self.table)

        # The control column has more groups than fit at a normal window
        # height; without this they compress until the spinboxes and their
        # labels overlap and cannot be clicked. A scroll area lets each group
        # keep its natural size and the user scroll to the rest.
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        right_scroll.setWidget(right)
        right_scroll.setMinimumWidth(340)

        splitter.addWidget(left)
        splitter.addWidget(right_scroll)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)

    def _toggle_manual_scale(self, checked):
        self.spin_bar_px.setEnabled(checked)
        self.spin_bar_real.setEnabled(checked)
        self.combo_unit.setEnabled(checked)
        self.btn_measure.setEnabled(checked and self.original_image is not None)
        if not checked:
            self.btn_measure.setChecked(False)
        if checked:
            self.scale_status.setText("수동 입력 모드")

    def _toggle_measure_mode(self, checked):
        self.image_label.measure_mode = checked
        self._measuring_shown = None
        self.image_label.setCursor(Qt.CrossCursor if checked else Qt.ArrowCursor)
        if not checked:
            # Keep the line that was just measured; only a half-finished
            # two-click measurement is stale here.
            self.image_label.clear_measurement(line=False)
            return
        self.statusBar().showMessage(
            "스케일바를 따라 드래그하거나, 양 끝을 한 번씩 클릭하세요 "
            "(클릭 사이에 확대·이동 가능). Shift를 누르면 수평/수직 고정.")

    def _on_measure_anchor(self, pending):
        if pending:
            self._measuring_shown = None
            self.statusBar().showMessage(
                "한쪽 끝을 찍었습니다. 확대(휠)·이동(가운데 버튼 드래그)한 뒤 "
                "반대쪽 끝을 클릭하세요. Esc로 취소.")

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape and self.image_label.measure_mode:
            self.image_label.clear_measurement()
            self.statusBar().showMessage("측정을 취소했습니다.")
            return
        super().keyPressEvent(event)

    def _on_zoom_changed(self, percent):
        self.lbl_zoom.setText(f"{percent:.0f}%")

    def _on_scalebar_measuring(self, start, end, locked):
        length = math.hypot(end[0] - start[0], end[1] - start[1])
        # Mouse-move fires far faster than anyone can read; only redraw the
        # status bar when the number it shows actually changes.
        if round(length, 1) == self._measuring_shown:
            return
        self._measuring_shown = round(length, 1)
        hint = "축 고정 (Shift)" if locked else "Shift를 누르면 수평/수직으로 고정"
        self.statusBar().showMessage(f"드래그 중: {length:.1f} px   —   {hint}")

    def _on_scalebar_measured(self, start, end):
        p0, p1 = start, end
        length = math.hypot(p1[0] - p0[0], p1[1] - p0[1])
        if length < 2:
            return

        # A couple of pixels of drag error multiply into every diameter, so
        # snap to the bar actually under the drag where one can be found.
        snapped = ParticleAnalyzer.snap_scalebar(self.original_image, p0, p1)
        real = self.spin_bar_real.value()
        unit = self.combo_unit.currentText()
        if snapped:
            self.spin_bar_px.setValue(snapped)
            note = f"드래그 {length:.1f} px → 스케일바에 맞춤 {snapped:.1f} px."
        else:
            self.spin_bar_px.setValue(length)
            note = (f"스케일바 길이 {length:.1f} px (드래그 값 그대로 사용 — "
                    "이미지에서 막대를 찾지 못했습니다).")
        self.btn_measure.setChecked(False)
        self.statusBar().showMessage(
            f"{note}  '실제 길이'가 {real:g} {unit}가 맞는지 확인하세요.")

    def _on_measure_drag(self, start, end):
        # One drag mechanism, two purposes: measuring the scale bar, or
        # measuring a particle's diameter for calibration. Which one depends on
        # the mode the user turned on.
        if self._calib_mode:
            self._on_calib_measured(start, end)
        else:
            self._on_scalebar_measured(start, end)

    def _toggle_calib_mode(self, on):
        # Calibration and scale-bar measuring share the drag, so only one can
        # be armed at a time.
        if on and self.btn_measure.isChecked():
            self.btn_measure.setChecked(False)
        self._calib_mode = on
        self.image_label.measure_mode = on or self.btn_measure.isChecked()
        self.image_label.setCursor(Qt.CrossCursor if on else Qt.ArrowCursor)
        self.btn_calib.setText("지름 재기 중지" if on else "지름 재기 시작")
        if on:
            self.statusBar().showMessage(
                "입자의 지름을 가로질러 드래그하세요. 몇 개 재고 '보정 적용'을 누르면 "
                "전체 입자가 그 기준으로 다시 측정됩니다.")

    def _on_calib_measured(self, start, end):
        length = math.hypot(end[0] - start[0], end[1] - start[1])
        if length < 3:
            return
        cx = (start[0] + end[0]) / 2.0
        cy = (start[1] + end[1]) / 2.0
        self._calib_refs.append((cx, cy, length))
        self.btn_calib_apply.setEnabled(len(self._calib_refs) >= 3)
        self.btn_calib_reset.setEnabled(True)
        u = self.unit
        dia = length * (self.nm_per_px or 1.0)
        self.lbl_calib.setText(f"측정 {len(self._calib_refs)}개 "
                               f"(마지막 {dia:.1f} {u})")
        self._redraw_results()
        need = max(0, 3 - len(self._calib_refs))
        tail = f" — {need}개 더 재면 적용할 수 있습니다" if need else " — '보정 적용' 준비됨"
        self.statusBar().showMessage(f"측정 {len(self._calib_refs)}개 수집{tail}.")

    def _apply_calibration(self):
        if not self.particles or len(self._calib_refs) < 3:
            return
        analyzer = ParticleAnalyzer(
            nm_per_px=self.nm_per_px,
            edge_level=("auto" if self.chk_edge_auto.isChecked()
                        else self.spin_edge.value() / 100.0),
            sphere_edge=self.chk_sphere_edge.isChecked())
        before = np.median([p["diameter"] for p in self._valid_particles()]) \
            if self._valid_particles() else 0.0
        # Only compute the place here; the re-analysis below applies it through
        # the full pipeline, so the number the user sees now is exactly what a
        # later re-run (or the same specimen at another magnification) gives.
        place, used = analyzer.calibrate_to_measurements(
            self.particles, self._calib_refs, self.original_image, apply=False)
        if place is None:
            self.statusBar().showMessage(
                "보정 실패: 측정한 위치에서 쉘 벽을 찾지 못했습니다. "
                "입자 경계가 뚜렷한 곳을 재보세요.")
            return
        self.wall_place = place
        self.btn_calib.setChecked(False)
        self._run_analysis()
        # Persist immediately, so the calibration survives a restart without the
        # user having to remember to save it.
        self._save_settings(config.default_path(), quiet=True)
        after = np.median([p["diameter"] for p in self._valid_particles()]) \
            if self._valid_particles() else 0.0
        self.lbl_calib.setText(f"적용됨 · 벽 위치 {place:.2f} (참조 {used}개)")
        u = self.unit
        self.statusBar().showMessage(
            f"보정 적용: 벽 위치 {place:.2f}로 전체 재측정. "
            f"D50 {before:.1f} → {after:.1f} {u}. "
            "같은 시료의 다른 배율 이미지에도 자동 적용됩니다.")

    def _reset_calibration(self):
        self._calib_refs = []
        self.wall_place = None
        self.btn_calib_apply.setEnabled(False)
        self.btn_calib_reset.setEnabled(False)
        self.lbl_calib.setText("측정 0개")
        if self.original_image is not None and self.particles:
            self._run_analysis()
        # Clearing is a deliberate act, so it clears the persisted file too;
        # otherwise the next launch would silently reapply the old calibration.
        self._save_settings(config.default_path(), quiet=True)
        self.statusBar().showMessage("보정을 초기화했습니다. 자동 판단으로 되돌립니다.")

    # -- settings persistence -------------------------------------------------

    def _settings_dict(self):
        """Gather the current configuration into a plain dict (no scale)."""
        return {
            "wall_place": self.wall_place,
            "min_area": self.spin_min_area.value(),
            "max_area": self.spin_max_area.value(),
            "circularity": self.spin_circularity.value(),
            "hollow": self.chk_hollow.isChecked(),
            "watershed": self.chk_watershed.isChecked(),
            "core": self.chk_core.isChecked(),
            "shell": self.chk_shell.isChecked(),
            "sphere_edge": self.chk_sphere_edge.isChecked(),
            "edge_auto": self.chk_edge_auto.isChecked(),
            "edge_level": self.spin_edge.value(),
        }

    def _apply_settings_dict(self, d):
        """Push a loaded settings dict onto the widgets and the calibration."""
        if "min_area" in d:
            self.spin_min_area.setValue(int(d["min_area"]))
        if "max_area" in d:
            self.spin_max_area.setValue(int(d["max_area"]))
        if "circularity" in d:
            self.spin_circularity.setValue(float(d["circularity"]))
        if "hollow" in d:
            self.chk_hollow.setChecked(bool(d["hollow"]))
        if "watershed" in d:
            self.chk_watershed.setChecked(bool(d["watershed"]))
        if "core" in d:
            self.chk_core.setChecked(bool(d["core"]))
        if "shell" in d:
            self.chk_shell.setChecked(bool(d["shell"]))
        if "sphere_edge" in d:
            self.chk_sphere_edge.setChecked(bool(d["sphere_edge"]))
        if "edge_auto" in d:
            self.chk_edge_auto.setChecked(bool(d["edge_auto"]))
        if "edge_level" in d:
            self.spin_edge.setValue(int(d["edge_level"]))
        if "wall_place" in d:
            wp = d["wall_place"]
            self.wall_place = float(wp) if wp is not None else None
            if self.wall_place is not None:
                self.lbl_calib.setText(f"불러옴 · 벽 위치 {self.wall_place:.2f}")
                self.btn_calib_reset.setEnabled(True)

    def _save_settings(self, path, quiet=False):
        try:
            config.save_settings(path, self._settings_dict())
        except OSError as e:
            if not quiet:
                self.statusBar().showMessage(f"설정 저장 실패: {e}")
            return False
        if not quiet:
            self.statusBar().showMessage(f"설정을 저장했습니다: {path}")
        return True

    def _load_and_apply_settings(self, path, quiet=False):
        d = config.load_settings(path)
        if not d:
            if not quiet:
                self.statusBar().showMessage("불러올 설정이 없습니다.")
            return False
        self._apply_settings_dict(d)
        return True

    def _save_settings_as(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "설정 저장", "tem_settings.json", "설정 파일 (*.json)")
        if path:
            self._save_settings(path)

    def _load_settings_from(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "설정 불러오기", "", "설정 파일 (*.json);;모든 파일 (*)")
        if not path:
            return
        if self._load_and_apply_settings(path):
            note = (f"벽 위치 {self.wall_place:.2f} 적용됨"
                    if self.wall_place is not None else "보정 없음")
            msg = f"설정을 불러왔습니다 ({note})."
            if self.original_image is not None and self.particles:
                self._run_analysis()
                msg += " 재분석했습니다."
            self.statusBar().showMessage(msg)

    # -- ground-truth labelling (archive only, never overrides measurement) ---

    def _current_labels(self):
        if self._image_key is None:
            return []
        return labelstore.labels_for(self._labels, self._image_key)

    def _refresh_label_count(self):
        n = len(self._current_labels())
        self.lbl_label_count.setText(f"이 이미지 라벨 {n}개")
        self.btn_label_report.setEnabled(n > 0 and bool(self.particles))
        self.btn_label_export.setEnabled(bool(self._labels))
        self.btn_label_calib.setEnabled(
            n >= 3 and bool(self.particles) and self.nm_per_px is not None)

    def _calibrate_from_labels(self):
        # The archived true values ARE calibration references: each is a
        # (position, true diameter), which is what the wall-place calibration
        # needs. Feeding them in is what turns "recording data" into "the
        # measurement getting better" - and because the resulting place is
        # stored like any calibration, the next image of the same specimen
        # picks it up without any labelling at all.
        labs = self._current_labels()
        if len(labs) < 3 or self.nm_per_px is None or not self.particles:
            return
        analyzer = ParticleAnalyzer(
            nm_per_px=self.nm_per_px,
            edge_level=("auto" if self.chk_edge_auto.isChecked()
                        else self.spin_edge.value() / 100.0),
            sphere_edge=self.chk_sphere_edge.isChecked())
        before = np.median([p["diameter"] for p in self._valid_particles()]) \
            if self._valid_particles() else 0.0
        # Each label already carries the place its wall gave (measured when it
        # was recorded), so the calibration comes straight from the stored
        # data - no image needed, and identical to what a machine holding only
        # labels.json would compute. Labels without a stored place (older ones)
        # fall back to measuring against the current image.
        stored = [lab["place"] for lab in labs if lab.get("place") is not None]
        if len(stored) >= 3:
            place, used = float(np.median(stored)), len(stored)
            analyzer.wall_place = place
        else:
            refs = [(lab["cx"], lab["cy"], lab["true_nm"] / self.nm_per_px)
                    for lab in labs]
            place, used = analyzer.calibrate_to_measurements(
                self.particles, refs, self.original_image, apply=False)
        if place is None:
            self.statusBar().showMessage(
                "보정 실패: 라벨 위치에서 쉘 벽을 찾지 못했습니다.")
            return
        self.wall_place = place
        self._run_analysis()
        self._save_settings(config.default_path(), quiet=True)
        after = np.median([p["diameter"] for p in self._valid_particles()]) \
            if self._valid_particles() else 0.0
        self.lbl_calib.setText(f"적용됨 · 벽 위치 {place:.2f} (라벨 {used}개)")
        self.btn_calib_reset.setEnabled(True)
        u = self.unit
        self.statusBar().showMessage(
            f"라벨 {used}개로 보정: 벽 위치 {place:.2f}. D50 {before:.1f} → {after:.1f} {u}. "
            "저장되어 같은 시료의 다음 이미지에 자동 적용됩니다.")

    def _toggle_label_mode(self, on):
        # Labelling clicks a particle; it must not fight the scale-bar or
        # calibration drag, so arming it disarms those.
        if on:
            if self.btn_measure.isChecked():
                self.btn_measure.setChecked(False)
            if self.btn_calib.isChecked():
                self.btn_calib.setChecked(False)
            self.image_label.measure_mode = False
        self._label_mode = on
        self.image_label.setCursor(Qt.PointingHandCursor if on else Qt.ArrowCursor)
        self.btn_label.setText("라벨 클릭 중지" if on else "라벨 클릭 시작")
        if on:
            self.statusBar().showMessage(
                "입자를 클릭해 참 지름을 입력하세요. 측정값은 그대로 두고 참값만 "
                "기록됩니다. 같은 곳 다시 클릭=수정, 0 입력=삭제.")

    def _label_click(self, ox, oy):
        if self.original_image is None:
            return
        # Snap to the nearest detection so the label sits on a particle centre,
        # but fall back to the click itself so a missed particle can still be
        # recorded. The matched detection's diameter is offered as a starting
        # value and stored beside the true one, so the error is kept with the
        # label rather than recomputed against a later, different analysis.
        best, best_d = None, None
        for p in self._valid_particles():
            d = np.hypot(p["center_x"] - ox, p["center_y"] - oy)
            if best_d is None or d < best_d:
                best, best_d = p, d
        prog_nm = None
        cx, cy = float(ox), float(oy)
        if best is not None and best_d <= best["radius_px"] * 1.2:
            cx, cy = float(best["center_x"]), float(best["center_y"])
            prog_nm = float(best["diameter"])

        existing = labelstore.nearest(self._current_labels(), cx, cy)
        near = (existing is not None
                and (existing["cx"] - cx) ** 2 + (existing["cy"] - cy) ** 2
                <= labelstore.REPLACE_RADIUS ** 2)
        u = self.unit
        preset = (existing["true_nm"] if near
                  else (prog_nm if prog_nm is not None else 0.0))
        value, ok = QInputDialog.getDouble(
            self, "참값 입력",
            (f"참 지름 ({u}):   [측정값 {prog_nm:.1f} {u}]" if prog_nm is not None
             else f"참 지름 ({u}):") + "\n(0 = 이 라벨 삭제)",
            float(preset), 0.0, 1e6, 2)
        if not ok:
            return
        if value <= 0:
            if labelstore.remove_near(self._labels, self._image_key, cx, cy):
                self._save_labels()
                self._draw_results()
                self._refresh_label_count()
                self.statusBar().showMessage("라벨을 삭제했습니다.")
            return
        # Measure the wall at this point and keep it with the label: the flank
        # positions, the place the user's diameter implies across them, and the
        # radial profile they came from. This is what makes the calibration
        # reproducible from the labels alone, without the image.
        extra = self._wall_evidence(cx, cy, best, value)
        _, replaced = labelstore.add_or_replace(
            self._labels, self._image_key, cx, cy, value, prog_nm,
            when=datetime.now().isoformat(timespec="seconds"), extra=extra)
        self._save_labels()
        self._draw_results()
        self._refresh_label_count()
        err = f"  (측정 {prog_nm:.1f} → 오차 {(prog_nm - value) / value * 100:+.1f}%)" \
            if prog_nm else ""
        self.statusBar().showMessage(
            (f"라벨 {'수정' if replaced else '기록'}: {value:.1f} {u}{err}. "
             f"이 이미지 {len(self._current_labels())}개."))

    def _wall_evidence(self, cx, cy, particle, true_nm):
        """Measure the wall at a labelled point, returning it plus the place.

        The wall context (blurred image, field wall fraction) is built once per
        analysis and cached, so labelling many particles does not re-scan the
        field each time.
        """
        if self._wall_ctx is None:
            analyzer = ParticleAnalyzer(
                nm_per_px=self.nm_per_px,
                edge_level=("auto" if self.chk_edge_auto.isChecked()
                            else self.spin_edge.value() / 100.0),
                sphere_edge=self.chk_sphere_edge.isChecked())
            self._wall_analyzer = analyzer
            self._wall_ctx = analyzer.wall_context(self.original_image, self.particles)
        if self._wall_ctx is None or particle is None:
            return None
        wall = self._wall_analyzer.measure_wall(
            cx, cy, particle["radius_px"], self._wall_ctx)
        if wall is None:
            return None
        # Where the user's true diameter sits across this wall (0 = inner flank,
        # 1 = outer). Stored so the calibration is derivable without the image.
        true_px = (true_nm / self.nm_per_px) / 2.0 if self.nm_per_px else None
        if true_px is not None and wall["outer_px"] > wall["inner_px"]:
            wall["place"] = float((true_px - wall["inner_px"])
                                  / (wall["outer_px"] - wall["inner_px"]))
        wall["nm_per_px"] = self.nm_per_px
        return wall

    def _save_labels(self):
        try:
            labelstore.save(labelstore.default_path(), self._labels)
        except OSError as e:
            self.statusBar().showMessage(f"라벨 저장 실패: {e}")

    def _label_errors(self):
        """Match each label on this image to the nearest current detection."""
        pairs = []
        valid = self._valid_particles()
        for lab in self._current_labels():
            best, best_d = None, None
            for p in valid:
                d = np.hypot(p["center_x"] - lab["cx"], p["center_y"] - lab["cy"])
                if best_d is None or d < best_d:
                    best, best_d = p, d
            if best is not None and best_d <= best["radius_px"]:
                pairs.append((lab["true_nm"], float(best["diameter"])))
        return pairs

    def _accuracy_report(self):
        pairs = self._label_errors()
        if not pairs:
            QMessageBox.information(
                self, "정확도 리포트",
                "이 이미지의 라벨에 대응하는 검출을 찾지 못했습니다.")
            return
        rel = np.array([(prog - true) / true * 100 for true, prog in pairs])
        u = self.unit
        QMessageBox.information(
            self, "정확도 리포트",
            f"라벨 {len(pairs)}개 기준 (이 이미지)\n\n"
            f"평균 절대 오차: {np.abs(rel).mean():.2f} %\n"
            f"편향(평균 오차): {rel.mean():+.2f} %\n"
            f"표준편차: {rel.std():.2f} %\n"
            f"최대 오차: {np.abs(rel).max():.2f} %\n\n"
            "※ 대표 입자를 골고루 라벨해야 실제 정확도에 가깝습니다. "
            "틀린 것만 라벨하면 나쁘게 나옵니다.")

    def _export_labels_csv(self):
        if not self._labels:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "라벨 CSV 내보내기", "tem_labels.csv", "CSV 파일 (*.csv)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                f.write("image,center_x,center_y,true_diameter,program_diameter,"
                        "error_pct,wall_inner_px,wall_outer_px,wall_place,time\n")
                for key, labs in self._labels.items():
                    name = key.split("|")[0]
                    for lab in labs:
                        prog = lab.get("prog_nm")
                        err = ("" if not prog else
                               f"{(prog - lab['true_nm']) / lab['true_nm'] * 100:.2f}")
                        f.write(f"{name},{lab['cx']},{lab['cy']},{lab['true_nm']},"
                                f"{'' if prog is None else prog},{err},"
                                f"{lab.get('inner_px', '')},{lab.get('outer_px', '')},"
                                f"{lab.get('place', '')},{lab.get('time') or ''}\n")
        except OSError as e:
            self.statusBar().showMessage(f"CSV 저장 실패: {e}")
            return
        total = sum(len(v) for v in self._labels.values())
        self.statusBar().showMessage(f"라벨 {total}개를 CSV로 저장했습니다: {path}")

    def _load_image(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "TEM 이미지 열기", "",
            "이미지 파일 (*.png *.jpg *.jpeg *.tif *.tiff *.bmp);;모든 파일 (*)"
        )
        if not path:
            return

        self.image = load_image(path)
        if self.image is None:
            QMessageBox.warning(self, "오류", "이미지를 읽을 수 없습니다.")
            return

        self.original_image = self.image.copy()
        self._image_path = path
        self._image_key = labelstore.image_key(path, self.original_image.shape)
        self.btn_label.setEnabled(True)
        self._refresh_label_count()

        if not self.chk_manual.isChecked():
            detector = ScaleBarDetector()
            nm_per_px, scale_text = detector.detect(self.image)
            if nm_per_px:
                self.nm_per_px = nm_per_px
                self.scale_text = scale_text
                self.scale_status.setText(f"감지됨: {scale_text} ({nm_per_px:.4f} nm/px)")
                self.statusBar().showMessage(f"스케일바 자동 감지: {scale_text}")
            else:
                self.scale_status.setText("자동 감지 실패 - 수동 입력 필요")
                self.chk_manual.setChecked(True)
                self.statusBar().showMessage("스케일바 자동 감지 실패. 수동으로 입력해주세요.")

        self._display_image(self.image, reset_view=True)
        self.btn_analyze.setEnabled(True)
        self.btn_measure.setEnabled(self.chk_manual.isChecked())
        self.statusBar().showMessage(
            f"이미지 로드 완료: {os.path.basename(path)}  —  "
            "'이미지에서 스케일바 재기'로 배율을 먼저 맞추세요.")

    def _run_analysis(self):
        if self.image is None:
            return
        # The field changed, so the cached wall context is stale.
        self._wall_ctx = None

        if self.chk_manual.isChecked():
            bar_px = self.spin_bar_px.value()
            bar_real = self.spin_bar_real.value()
            unit = self.combo_unit.currentText()
            unit_to_nm = {"nm": 1, "μm": 1000, "mm": 1e6}
            self.nm_per_px = (bar_real * unit_to_nm[unit]) / bar_px
            self.unit = "nm"

        max_area = self.spin_max_area.value()
        if max_area == 0:
            max_area = None

        # Analysis runs on the UI thread and takes seconds on full-resolution
        # TEM images, so make the wait visible instead of looking frozen. The
        # analyzer reports progress through a callback; because we are on the UI
        # thread, processEvents lets the bar actually repaint between stages.
        self.statusBar().showMessage("분석 중...")
        self.btn_analyze.setEnabled(False)
        self.progress.setValue(0)
        self.progress.setFormat("분석 준비 중… %p%")
        self.progress.setVisible(True)
        QApplication.setOverrideCursor(Qt.WaitCursor)
        QApplication.processEvents()

        def _on_progress(frac, label):
            self.progress.setValue(int(frac * 100))
            self.progress.setFormat(f"{label}… %p%")
            QApplication.processEvents()

        try:
            analyzer = ParticleAnalyzer(
                nm_per_px=self.nm_per_px,
                edge_level=("auto" if self.chk_edge_auto.isChecked()
                            else self.spin_edge.value() / 100.0),
                sphere_edge=self.chk_sphere_edge.isChecked())
            # Carry a hand calibration, if one has been set, into this run - so
            # re-analysing, or opening the same specimen at another
            # magnification, keeps the edge where the user put it.
            analyzer.wall_place = self.wall_place
            self.particles = analyzer.analyze(
                self.original_image,
                min_area_px=self.spin_min_area.value(),
                max_area_px=max_area,
                circularity_thresh=self.spin_circularity.value(),
                use_watershed=self.chk_watershed.isChecked(),
                hollow=self.chk_hollow.isChecked(),
                    detect_cores=self.chk_core.isChecked(),
                measure_shell=self.chk_shell.isChecked(),
                progress=_on_progress,
            )
        finally:
            QApplication.restoreOverrideCursor()
            self.btn_analyze.setEnabled(True)
            self.progress.setVisible(False)

        self._draw_results()
        stats = ParticleAnalyzer.compute_statistics(self.particles)
        self._update_stats(stats)
        self._update_table()
        self._update_histogram()
        valid = self._valid_particles()
        self.btn_export.setEnabled(bool(valid))
        self.btn_calib.setEnabled(bool(valid))
        self._refresh_label_count()
        n_exc = len(self.particles) - len(valid)
        n_approx = sum(1 for p in valid if p.get("approx"))
        msg = f"분석 완료: {len(valid)}개 입자 검출"
        if n_approx:
            msg += f" (근사 {n_approx}개 포함)"
        if n_exc:
            msg += f", 판별불가 {n_exc}개 제외"
        n_aside = sum(1 for p in self.particles
                      if p.get("unoutlined") and not p.get("restored"))
        if n_aside:
            msg += (f"  |  확인 필요 {n_aside}개 (주황색) - 둘레가 덜 보여 "
                    "유령과 구분되지 않음, 클릭하면 포함")
        n_defect = sum(1 for p in valid if p.get("defect"))
        if n_defect:
            msg += (f"  |  불량 {n_defect}개 (파란 안쪽 원) - 코어 잔존 또는 "
                    "내부 파손, 크기는 측정됨")
        n_irreg = sum(1 for p in valid if p.get("irregular"))
        if n_irreg:
            msg += f"  |  비원형 {n_irreg}개 (자홍색 번호) - 길쭉하거나 뭉친 입자"
        n_overlap, _ = self._overlapping_pairs()
        if n_overlap:
            msg += f"  |  겹치는 검출 {n_overlap}쌍 (보라색) - 중복 여부 확인 필요"
        self.statusBar().showMessage(msg)

    def _redraw_results(self):
        """Re-render the annotated view without re-running the analysis."""
        if self.particles and self.original_image is not None:
            self._draw_results()

    def _valid_particles(self):
        return [p for p in self.particles if not p.get("excluded")]

    def _overlapping_pairs(self):
        """Valid detections that sit substantially on top of one another.

        Two circles covering the same particle inflate the count without
        looking obviously wrong in a crowded field, so they are called out
        rather than left to be spotted by eye.

        The test is how much of the *smaller* circle the overlap eats. Comparing
        centre distance against the summed radii misses the case that matters
        most - a small detection sitting almost entirely inside a large one -
        because the large radius alone keeps the threshold out of reach.
        """
        valid = self._valid_particles()
        flagged = set()
        pairs = 0
        for i, p in enumerate(valid):
            for j in range(i + 1, len(valid)):
                q = valid[j]
                d = np.hypot(p["center_x"] - q["center_x"],
                             p["center_y"] - q["center_y"])
                if ParticleAnalyzer.overlap_fraction(d, p["radius_px"], q["radius_px"]) > 0.30:
                    pairs += 1
                    flagged.add(id(p))
                    flagged.add(id(q))
        return pairs, flagged

    def _draw_results(self):
        h, w = self.original_image.shape[:2]
        scale = max(1, int(round(900 / max(h, w))))
        self._result_scale = scale
        display = cv2.resize(self.original_image, None, fx=scale, fy=scale,
                             interpolation=cv2.INTER_CUBIC) if scale > 1 else self.original_image.copy()
        thickness = max(2, scale)
        _, overlapping = self._overlapping_pairs()
        num = 0
        for p in self.particles:
            cx, cy = p["center_x"] * scale, p["center_y"] * scale
            r = int(p["radius_px"]) * scale
            excluded = p.get("excluded", False)
            if excluded and p.get("unoutlined") and not p.get("restored"):
                # Not deleted, offered. The measurement that finds phantoms -
                # how much of the circumference carries a dark ring - cannot
                # tell a phantom from a real particle whose neighbours hide
                # most of its outline; the two overlap completely on every
                # figure measured. So these are set aside rather than thrown
                # away, and a click puts one back.
                color = (60, 170, 255)
            elif excluded:
                color = (0, 0, 255)
            elif p.get("overlap") or id(p) in overlapping:
                color = (255, 0, 255)
            elif p.get("approx"):
                color = (0, 220, 220)
            else:
                color = (0, 220, 0)
            if p.get("contour") is not None:
                # Scale first, then round: the outline is sub-pixel, and rounding
                # it to whole source pixels before the upscale turns every
                # half-pixel wobble into a `scale`-pixel staircase.
                scaled_cnt = np.round(p["contour"].astype(np.float64) * scale).astype(np.int32)
                # A particle the frame cuts is drawn as the arc that was
                # actually seen, with the rest a hairline, whatever the option
                # says. Its diameter comes from that arc and the closing
                # stretch is an extrapolation; drawn at full weight it reads as
                # a measured boundary lying over blank frame, and a reader
                # rightly marks it as an error.
                self._draw_boundary(display, scaled_cnt, (cx, cy), color, thickness,
                                    p.get("contour_measured"),
                                    self.chk_mark_inferred.isChecked()
                                    or bool(p.get("partial")))
            else:
                cv2.circle(display, (cx, cy), r, color, thickness)
            if excluded:
                continue
            num += 1
            if p.get("inner_radius_px"):
                cv2.circle(display, (cx, cy), int(p["inner_radius_px"] * scale),
                           (255, 200, 0), max(1, thickness - 1))
            if p.get("has_core"):
                cv2.drawMarker(display, (cx, cy), (0, 165, 255),
                               cv2.MARKER_CROSS, max(10, r // 3), thickness)
            if p.get("defect"):
                # Measured and counted like any other particle - the ring is a
                # label, not a verdict on the measurement - so it goes inside
                # the boundary rather than replacing its colour, which still
                # says how well the diameter is known.
                cv2.circle(display, (cx, cy), max(2, int(r * 0.82)),
                           (255, 60, 60), max(1, thickness - 1))
            txt = str(num)
            pos = (cx - 8 * len(txt), cy - 6)
            font_scale = 0.45 + 0.1 * scale
            # A number in magenta says the boundary is not a circle - an
            # elongated particle, or two fused - so its diameter means less
            # than the others'. It goes on the number rather than the outline
            # because the outline colour is already saying how well the edge
            # was found, which is a different question.
            ink = (255, 80, 255) if p.get("irregular") else (80, 255, 255)
            cv2.putText(display, txt, pos, cv2.FONT_HERSHEY_SIMPLEX,
                        font_scale, (0, 0, 0), 3)
            cv2.putText(display, txt, pos, cv2.FONT_HERSHEY_SIMPLEX,
                        font_scale, ink, 1)

        legend = [("OK", (0, 220, 0)), ("Approx", (0, 220, 220)),
                  ("Overlap", (255, 0, 255)), ("Excluded", (0, 0, 255))]
        lx = 10
        for label, color in legend:
            cv2.rectangle(display, (lx, 10), (lx + 18, 28), color, -1)
            cv2.putText(display, label, (lx + 24, 26), cv2.FONT_HERSHEY_SIMPLEX,
                        0.6, (0, 0, 0), 3)
            cv2.putText(display, label, (lx + 24, 26), cv2.FONT_HERSHEY_SIMPLEX,
                        0.6, (255, 255, 255), 1)
            lx += 24 + 13 * len(label) + 16

        # The diameters measured for calibration, drawn as cyan lines through
        # their centres so the user can see what they have collected so far.
        for cx0, cy0, length in self._calib_refs:
            c = (int(cx0 * scale), int(cy0 * scale))
            half = int(length * scale / 2)
            cv2.line(display, (c[0] - half, c[1]), (c[0] + half, c[1]),
                     (255, 255, 0), max(1, thickness))
            cv2.circle(display, c, max(2, thickness), (255, 255, 0), -1)

        # Ground-truth labels: a magenta marker where a true diameter was
        # recorded, with the value and, when it matches a detection, the error.
        # It shows the user which particles they have checked without touching
        # the measurement.
        valid = self._valid_particles()
        for lab in self._current_labels():
            c = (int(lab["cx"] * scale), int(lab["cy"] * scale))
            s = max(3, 2 * thickness)
            cv2.drawMarker(display, c, (255, 0, 255), cv2.MARKER_SQUARE, s * 2, thickness)
            txt = f"{lab['true_nm']:.0f}"
            best, best_d = None, None
            for p in valid:
                d = np.hypot(p["center_x"] - lab["cx"], p["center_y"] - lab["cy"])
                if best_d is None or d < best_d:
                    best, best_d = p, d
            if best is not None and best_d <= best["radius_px"]:
                txt += f" ({(best['diameter'] - lab['true_nm']) / lab['true_nm'] * 100:+.0f}%)"
            org = (c[0] + s + 2, c[1] + s // 2)
            cv2.putText(display, txt, org, cv2.FONT_HERSHEY_SIMPLEX,
                        0.4 + 0.06 * scale, (0, 0, 0), thickness + 2)
            cv2.putText(display, txt, org, cv2.FONT_HERSHEY_SIMPLEX,
                        0.4 + 0.06 * scale, (255, 0, 255), thickness)

        self.result_image = display
        self.btn_save_image.setEnabled(True)
        self._display_image(display)

    @staticmethod
    def _draw_boundary(display, contour, center, color, thickness, measured=None,
                       mark_inferred=False):
        """Draw the closed outline round the particle.

        By default the whole boundary is drawn at one weight, because a
        boundary drawn in pieces is not a boundary. With ``mark_inferred`` the
        stretches whose edge was never found - the contact sides, where a
        neighbour is pressed against this particle - are drawn as a hairline
        instead, which shows how much of the outline rests on a measurement at
        the cost of a busier picture.
        """
        pts = contour.reshape(-1, 2)
        if len(pts) < 2:
            return
        closed = np.vstack([pts, pts[:1]]).reshape(-1, 1, 2).astype(np.int32)

        if measured is None or len(measured) != len(pts) or not mark_inferred:
            cv2.polylines(display, [closed], False, color, thickness)
            return

        cv2.polylines(display, [closed], False, color, max(1, thickness - 2))
        runs, run = [], []
        for i, on in enumerate(measured):
            if on:
                run.append(pts[i])
            elif run:
                runs.append(run)
                run = []
        if run:
            if runs and measured[0]:
                runs[0] = run + runs[0]   # the outline wraps
            else:
                runs.append(run)
        for seg in runs:
            if len(seg) >= 2:
                cv2.polylines(display, [np.array(seg, np.int32).reshape(-1, 1, 2)],
                              False, color, thickness)

    def _on_image_click(self, ox, oy):
        # Labelling takes the click before anything else: in this mode a click
        # records a true diameter, it does not toggle a core or restore a
        # set-aside particle.
        if self._label_mode:
            self._label_click(ox, oy)
            return
        if not self.particles:
            return
        # A set-aside particle is put back, or a restored one set aside again.
        # This takes precedence over the core toggle: those circles are drawn
        # in their own colour and clicking one can only mean the one thing.
        aside, aside_d = None, None
        for p in self.particles:
            if not p.get("unoutlined"):
                continue
            d = np.hypot(p["center_x"] - ox, p["center_y"] - oy)
            if d <= p["radius_px"] and (aside_d is None or d < aside_d):
                aside, aside_d = p, d
        if aside is not None:
            restored = not aside.get("restored", False)
            aside["restored"] = restored
            aside["excluded"] = not restored
            self._draw_results()
            self._update_stats(ParticleAnalyzer.compute_statistics(self.particles))
            self.statusBar().showMessage(
                "확인 필요 입자를 " + ("포함했습니다" if restored else "다시 제외했습니다"))
            return

        if "has_core" not in (self.particles[0] if self.particles else {}):
            return
        best = None
        best_d = None
        for p in self.particles:
            if p.get("excluded"):
                continue
            d = np.hypot(p["center_x"] - ox, p["center_y"] - oy)
            if d <= p["radius_px"] and (best_d is None or d < best_d):
                best = p
                best_d = d
        if best is None:
            return
        best["has_core"] = not best["has_core"]
        self._draw_results()
        stats = ParticleAnalyzer.compute_statistics(self.particles)
        self._update_stats(stats)
        self._update_table()
        state = "체크" if best["has_core"] else "해제"
        self.statusBar().showMessage(f"코어 {state}: ({best['center_x']}, {best['center_y']}) 입자")

    def _save_result_image(self):
        if self.result_image is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "결과 이미지 저장", "particle_result.png",
            "PNG 이미지 (*.png);;JPEG 이미지 (*.jpg)"
        )
        if not path:
            return
        if not save_image(path, self.result_image):
            QMessageBox.warning(self, "오류", "이미지를 저장할 수 없습니다.")
            return
        self.statusBar().showMessage(f"결과 이미지 저장 완료: {path}")

    def _display_image(self, cv_img, reset_view=False):
        # The result view is rendered at an integer upscale; the view maps its
        # own coordinates back to original image pixels through this factor.
        if self.original_image is not None:
            source_scale = cv_img.shape[1] / self.original_image.shape[1]
        else:
            source_scale = 1.0
        if len(cv_img.shape) == 2:
            h, w = cv_img.shape
            qimg = QImage(cv_img.data, w, h, w, QImage.Format_Grayscale8)
        else:
            rgb = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb.shape
            qimg = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888)
        self.image_label.set_image(QPixmap.fromImage(qimg), reset_view=reset_view,
                                   source_scale=source_scale)

    def _update_stats(self, stats):
        if not stats:
            return
        u = self.unit
        self.lbl_count.setText(str(stats["count"]))
        self.lbl_mean.setText(f"{stats['mean']:.2f} {u}")
        self.lbl_std.setText(f"{stats['std']:.2f} {u}")
        self.lbl_min.setText(f"{stats['min']:.2f} {u}")
        self.lbl_max.setText(f"{stats['max']:.2f} {u}")
        self.lbl_d10.setText(f"{stats['d10']:.2f} {u}")
        self.lbl_d50.setText(f"{stats['d50']:.2f} {u}")
        self.lbl_d90.setText(f"{stats['d90']:.2f} {u}")
        if "core_ratio" in stats:
            self.lbl_core.setText(
                f"{stats['core_count']}/{stats['count']} ({stats['core_ratio']*100:.1f}%)")
        else:
            self.lbl_core.setText("-")
        if stats.get("defect_count"):
            self.lbl_defect.setText(
                f"{stats['defect_count']}/{stats['count']} "
                f"({stats['defect_ratio'] * 100:.1f}%)")
        else:
            self.lbl_defect.setText("0")
        if stats.get("irregular_count"):
            self.lbl_irregular.setText(
                f"{stats['irregular_count']}/{stats['count']} "
                f"({stats['irregular_ratio'] * 100:.1f}%)")
        else:
            self.lbl_irregular.setText("0")
        if "shell_mean" in stats:
            self.lbl_shell.setText(
                f"{stats['shell_mean']:.2f} ± {stats['shell_std']:.2f} {u}"
                f"  ({stats['shell_count']}개)")
            self.lbl_porosity.setText(f"{stats['porosity_mean'] * 100:.1f} %")
        else:
            self.lbl_shell.setText("-")
            self.lbl_porosity.setText("-")

    @staticmethod
    def _numeric_item(value, text):
        item = QTableWidgetItem()
        item.setData(Qt.DisplayRole, float(f"{value:.2f}"))
        item.setToolTip(text)
        return item

    def _update_table(self):
        particles = self._valid_particles()
        has_core_col = bool(particles) and "has_core" in particles[0]
        # Pixel diameter is shown next to the calibrated value so a mismatch
        # with hand measurements can be pinned on the scale or on the sizing.
        has_shell_col = any(p.get("shell_thickness") is not None for p in particles)
        headers = ["#", "직경", "직경(px)", "면적"]
        if has_shell_col:
            headers += ["쉘 두께", "공극률(%)"]
        if has_core_col:
            headers += ["코어"]
        has_defect_col = any(p.get("defect") for p in particles)
        if has_defect_col:
            headers += ["불량"]
        has_irregular_col = any(p.get("irregular") for p in particles)
        if has_irregular_col:
            headers += ["비원형"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(particles))
        u = self.unit
        for i, p in enumerate(particles):
            num = str(i + 1) + (" (근사)" if p.get("approx") else "")
            self.table.setItem(i, 0, self._numeric_item(i + 1, num))
            self.table.setItem(i, 1, self._numeric_item(p["diameter"], f"{p['diameter']:.2f} {u}"))
            self.table.setItem(i, 2, self._numeric_item(p["radius_px"] * 2, "px"))
            self.table.setItem(i, 3, self._numeric_item(p["area"], f"{u}²"))
            col = 4
            if has_shell_col:
                if p.get("shell_thickness") is not None:
                    self.table.setItem(i, col, self._numeric_item(p["shell_thickness"], u))
                    self.table.setItem(i, col + 1,
                                       self._numeric_item(p["porosity"] * 100, "%"))
                else:
                    self.table.setItem(i, col, QTableWidgetItem("-"))
                    self.table.setItem(i, col + 1, QTableWidgetItem("-"))
                col += 2
            if has_core_col:
                self.table.setItem(i, col, QTableWidgetItem("유" if p["has_core"] else "무"))
                col += 1
            if has_defect_col:
                self.table.setItem(i, col,
                                   QTableWidgetItem("불량" if p.get("defect") else ""))
                col += 1
            if has_irregular_col:
                self.table.setItem(i, col,
                                   QTableWidgetItem("비원형" if p.get("irregular") else ""))
        self.table.setSortingEnabled(True)

    def _update_histogram(self):
        diameters = [p["diameter"] for p in self._valid_particles()]
        self.histogram.plot(diameters, self.unit)

    def _export_excel(self):
        if not self.particles:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Excel로 내보내기", "particle_analysis.xlsx",
            "Excel 파일 (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Particle Data"

        u = self.unit
        particles = self._valid_particles()
        has_core = bool(particles) and "has_core" in particles[0]
        headers = ["#", f"Diameter ({u})", "Diameter (px)", f"Area ({u}²)",
                   "Center X (px)", "Center Y (px)"]
        headers += ["Approx"]
        has_shell = any(p.get("shell_thickness") is not None for p in particles)
        if has_shell:
            headers += [f"Inner Diameter ({u})", f"Shell Thickness ({u})", "Porosity"]
        if has_core:
            headers += ["Has Core"]
        headers += ["Defect", "Irregular"]
        ws_data.append(headers)
        for i, p in enumerate(particles):
            row = [i + 1, p["diameter"], p["radius_px"] * 2, p["area"],
                   p["center_x"], p["center_y"]]
            row += ["Y" if p.get("approx") else "N"]
            if has_shell:
                row += [p.get("inner_diameter"), p.get("shell_thickness"),
                        p.get("porosity")]
            if has_core:
                row += ["Y" if p["has_core"] else "N"]
            row += ["Y" if p.get("defect") else "N",
                    "Y" if p.get("irregular") else "N"]
            ws_data.append(row)

        ws_stats = wb.create_sheet("Statistics")
        stats = ParticleAnalyzer.compute_statistics(self.particles)
        ws_stats.append(["Metric", "Value", "Unit"])
        ws_stats.append(["Count", stats["count"], ""])
        ws_stats.append(["Mean Diameter", stats["mean"], u])
        ws_stats.append(["Std Dev", stats["std"], u])
        ws_stats.append(["Min", stats["min"], u])
        ws_stats.append(["Max", stats["max"], u])
        ws_stats.append(["D10", stats["d10"], u])
        ws_stats.append(["D50", stats["d50"], u])
        ws_stats.append(["D90", stats["d90"], u])
        if "shell_mean" in stats:
            ws_stats.append(["Shell Count", stats["shell_count"], ""])
            ws_stats.append(["Mean Shell Thickness", stats["shell_mean"], u])
            ws_stats.append(["Shell Thickness Std", stats["shell_std"], u])
            ws_stats.append(["Mean Inner Diameter", stats["inner_mean"], u])
            ws_stats.append(["Mean Porosity", stats["porosity_mean"], ""])
        if "core_ratio" in stats:
            ws_stats.append(["Core Count", stats["core_count"], ""])
            ws_stats.append(["Core Ratio", stats["core_ratio"], ""])
        if "defect_count" in stats:
            ws_stats.append(["Defect Count", stats["defect_count"], ""])
            ws_stats.append(["Defect Ratio", stats["defect_ratio"], ""])
        if "irregular_count" in stats:
            ws_stats.append(["Irregular Count", stats["irregular_count"], ""])
            ws_stats.append(["Irregular Ratio", stats["irregular_ratio"], ""])

        if self.nm_per_px:
            ws_stats.append([])
            ws_stats.append(["Scale", f"{self.nm_per_px:.4f}", "nm/px"])

        if self.result_image is not None:
            try:
                import tempfile
                from openpyxl.drawing.image import Image as XLImage
                # Close the handle before writing: Windows will not let the
                # image be written while the temp file is still open.
                tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                tmp.close()
                if save_image(tmp.name, self.result_image):
                    ws_img = wb.create_sheet("Result Image")
                    ws_img.add_image(XLImage(tmp.name), "A1")
            except Exception:
                pass

        wb.save(path)
        self.statusBar().showMessage(f"Excel 저장 완료: {path}")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
